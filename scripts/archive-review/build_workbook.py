import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

light  = json.load(open('rows_light.json'))
fading = json.load(open('rows_fading.json'))
rich   = json.load(open('rows_rich.json'))

rows = []
for r in light:
    rows.append(dict(r, summary=None, implication=None, category=None,
                     sources=None, stories=None, day_span=None, nearest=None,
                     shift_status_now=None, shift_momentum_now=None))
for f in fading:
    rows.append({'id': f[0], 'ref': f[1], 'title': f[2], 'week': '2026-W31',
                 'kind': 'shift_fading', 'created': '2026-07-31',
                 'summary': None, 'implication': None, 'category': f[6],
                 'sources': None, 'stories': None, 'day_span': None, 'nearest': None,
                 'last_evidence_on': f[3], 'shift_status_now': f[4], 'shift_momentum_now': f[5]})
rows += rich

KIND_LABEL = {'brief_review': 'Weekly brief', 'shift_proposal': 'New shift spotted',
              'shift_fading': 'Shift going quiet', 'graduation': 'Keep for good?',
              'investigation': 'Investigation'}
KIND_ORDER = ['brief_review', 'shift_proposal', 'shift_fading', 'graduation', 'investigation']
rows.sort(key=lambda r: (KIND_ORDER.index(r['kind']), r['week'], r['title']))

# Row numbers, then cross-reference duplicates by title and by shared ref.
for i, r in enumerate(rows):
    r['n'] = i + 1
by_title = {r['title']: r['n'] for r in rows}
ref_rows = {}
for r in rows:
    ref_rows.setdefault((r['kind'], r['ref']), []).append(r['n'])

for r in rows:
    notes = []
    same = [n for n in ref_rows[(r['kind'], r['ref'])] if n != r['n']]
    if same:
        notes.append('Same item as row ' + ', '.join(str(n) for n in same) + ' - rule once, I apply it to all')
    near = r.get('nearest')
    if near and near.get('similarity', 0) >= 15:
        tgt = by_title.get(near['title'])
        notes.append('Looks like a repeat of "%s"%s (%d%% alike)'
                     % (near['title'], ' - row %d' % tgt if tgt else '', near['similarity']))
    st = r.get('shift_status_now')
    if r['kind'] == 'shift_fading' and st == 'retired':
        notes.append('Already retired since - your call just confirms it')
    elif r['kind'] == 'shift_proposal' and st == 'retired':
        notes.append('This shift was retired later on')
    elif r['kind'] == 'shift_proposal' and st == 'proposed':
        notes.append('Still sitting unconfirmed in the register')
    r['heads_up'] = '. '.join(notes)

CALLS = {
    'brief_review':   ['Good brief', 'Not for me', 'Leave it'],
    'shift_proposal': ['Track this shift', 'Same as another one', 'Not a shift', 'Leave it'],
    'shift_fading':   ['Yes, retire it', 'No, still live', 'Leave it'],
    'graduation':     ['Keep for good', 'Not worth keeping', 'Leave it'],
    'investigation':  ['Got it', 'Not for me', 'Leave it'],
}
REASONS = ['Wrong topic for me', 'Wrong angle', 'Already covered', 'Old news',
           'Too generic', 'Not my voice', 'Bad timing', 'Other']

ARIAL   = 'Arial'
HEAD_FILL  = PatternFill('solid', fgColor='1F3864')
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')
DUP_FILL   = PatternFill('solid', fgColor='FCE4D6')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# ---------------------------------------------------------------- Lists sheet
ls = wb.create_sheet('Lists')
ls['A1'] = 'Dropdown sources. Do not edit - the Decide sheet points at these.'
ls['A1'].font = Font(ARIAL, bold=True)
col = 1
ranges = {}
for kind in KIND_ORDER:
    c = get_column_letter(col)
    ls.cell(row=2, column=col, value=KIND_LABEL[kind]).font = Font(ARIAL, bold=True)
    for j, v in enumerate(CALLS[kind]):
        ls.cell(row=3 + j, column=col, value=v).font = Font(ARIAL)
    ranges[kind] = "'Lists'!$%s$3:$%s$%d" % (c, c, 2 + len(CALLS[kind]))
    col += 1
c = get_column_letter(col)
ls.cell(row=2, column=col, value='Reason (if not for me)').font = Font(ARIAL, bold=True)
for j, v in enumerate(REASONS):
    ls.cell(row=3 + j, column=col, value=v).font = Font(ARIAL)
reason_range = "'Lists'!$%s$3:$%s$%d" % (c, c, 2 + len(REASONS))
for i in range(1, col + 1):
    ls.column_dimensions[get_column_letter(i)].width = 24

# ---------------------------------------------------------------- Decide sheet
ws = wb.create_sheet('Decide', 0)
HEADERS = ['#', 'Week', 'Card type', 'Title', 'What it found', 'So what',
           'Area', 'Evidence', 'Heads up',
           'YOUR CALL', 'WHY (only if not for me)', 'NOTES (optional)', 'SAME AS ROW #',
           'decision_id']
WIDTHS  = [5, 10, 19, 52, 74, 62, 14, 22, 46, 22, 24, 34, 14, 38]

ws['A1'] = 'Content decisions that aged out unseen - 64 cards, 2026-W28 to 2026-W33'
ws['A1'].font = Font(ARIAL, size=14, bold=True, color='1F3864')
ws['A2'] = ('Fill in the four shaded columns only. YOUR CALL is a dropdown and its options change by card type. '
            'Everything left blank stays archived and teaches nothing, which is the honest default. '
            'Row 4 is a filled-in example - delete it or leave it, it is ignored on import.')
ws['A2'].font = Font(ARIAL, size=10, italic=True, color='595959')
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
ws.row_dimensions[2].height = 30

HEAD_ROW = 3
for i, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=HEAD_ROW, column=i, value=h)
    cell.font = Font(ARIAL, bold=True, color='FFFFFF')
    cell.fill = HEAD_FILL
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = WIDTHS[i - 1]
ws.row_dimensions[HEAD_ROW].height = 30

EXAMPLE = ['e.g.', '2026-W31', 'New shift spotted', 'Example row - delete me',
           'What the engine found', 'What it means for you', 'economics', '4 sources, 6 stories',
           '', 'Not a shift', 'Too generic', 'Two data points is not a trend', '', '']
for i, v in enumerate(EXAMPLE, start=1):
    c2 = ws.cell(row=4, column=i, value=v)
    c2.font = Font(ARIAL, italic=True, color='808080')
    c2.alignment = Alignment(vertical='top', wrap_text=True)
    c2.border = BORDER
ws.row_dimensions[4].height = 28

FIRST = 5
for idx, r in enumerate(rows):
    rn = FIRST + idx
    ev = []
    if r.get('sources'):  ev.append('%s sources' % r['sources'])
    if r.get('stories'):  ev.append('%s stories' % r['stories'])
    if r.get('day_span'): ev.append('over %s days' % r['day_span'])
    if r.get('headlines'): ev.append('%s headlines' % r['headlines'])
    if r.get('last_evidence_on'): ev.append('last seen %s' % r['last_evidence_on'])
    if r.get('idea_state'): ev.append('idea is %s' % r['idea_state'])
    if r.get('terminal_rung'): ev.append('stopped at rung %s' % r['terminal_rung'])
    if r.get('shift_momentum_now') is not None: ev.append('momentum %s' % r['shift_momentum_now'])

    vals = [r['n'], r['week'], KIND_LABEL[r['kind']], r['title'],
            r.get('summary') or '', r.get('implication') or '',
            r.get('category') or '', ', '.join(ev), r['heads_up'],
            '', '', '', '', r['id']]
    for i, v in enumerate(vals, start=1):
        c2 = ws.cell(row=rn, column=i, value=v)
        c2.font = Font(ARIAL, size=10)
        c2.alignment = Alignment(vertical='top', wrap_text=(i in (4, 5, 6, 8, 9)))
        c2.border = BORDER
    for i in (10, 11, 12, 13):
        ws.cell(row=rn, column=i).fill = INPUT_FILL
    if r['heads_up']:
        ws.cell(row=rn, column=9).fill = DUP_FILL
    ws.cell(row=rn, column=14).font = Font(ARIAL, size=8, color='A6A6A6')
    ws.row_dimensions[rn].height = 62 if r.get('summary') else 30

# Per-kind dropdowns: the options must match the card, or the sheet invites a
# call the API cannot carry out.
for kind in KIND_ORDER:
    rn_list = [FIRST + i for i, r in enumerate(rows) if r['kind'] == kind]
    if not rn_list:
        continue
    dv = DataValidation(type='list', formula1=ranges[kind], allow_blank=True, showDropDown=False)
    dv.showErrorMessage = True
    dv.error = 'Pick one of the options offered for this card type.'
    dv.errorTitle = 'Not a valid call for this card'
    ws.add_data_validation(dv)
    for rn in rn_list:
        dv.add(ws.cell(row=rn, column=10))

dvr = DataValidation(type='list', formula1=reason_range, allow_blank=True, showDropDown=False)
dvr.showErrorMessage = True
dvr.error = 'Pick a reason from the list, or leave blank.'
dvr.errorTitle = 'Unknown reason'
ws.add_data_validation(dvr)
for idx in range(len(rows)):
    dvr.add(ws.cell(row=FIRST + idx, column=11))

ws.freeze_panes = 'D5'
ws.auto_filter.ref = 'A%d:M%d' % (HEAD_ROW, FIRST + len(rows) - 1)
ws.column_dimensions['N'].hidden = True

# ---------------------------------------------------------------- Read me
rm = wb.create_sheet('Read me first', 0)
rm.column_dimensions['A'].width = 30
rm.column_dimensions['B'].width = 104
LAST = FIRST + len(rows) - 1

def line(row, a, b, bold=False, size=11, color='000000'):
    rm.cell(row=row, column=1, value=a).font = Font(ARIAL, bold=True, size=size, color=color)
    c2 = rm.cell(row=row, column=2, value=b)
    c2.font = Font(ARIAL, size=size, bold=bold)
    c2.alignment = Alignment(wrap_text=True, vertical='top')
    rm.row_dimensions[row].height = 30

rm['A1'] = 'The 64 cards you never saw'
rm['A1'].font = Font(ARIAL, size=16, bold=True, color='1F3864')
rm['A2'] = ('These aged out of the Content queue between 10 July and 14 August without ever reaching you, '
            'because the queue was showing the oldest 30 of 74 and nothing ever cleared them. '
            'Nothing was deleted. Rule on them here and the learning loop finally gets fed.')
rm['A2'].font = Font(ARIAL, size=10, italic=True, color='595959')
rm.merge_cells('A2:B2'); rm.row_dimensions[2].height = 46

line(4, 'How many are done', '', size=12)
rm['B4'] = '=COUNTA(Decide!J%d:J%d)&" of %d ruled on, "&COUNTBLANK(Decide!J%d:J%d)&" still blank"' % (FIRST, LAST, len(rows), FIRST, LAST)
rm['B4'].font = Font(ARIAL, size=12, bold=True, color='1F3864')

line(6, 'Fill in these only', 'YOUR CALL, WHY, NOTES and SAME AS ROW # - the four shaded columns. Everything else is there so you can judge without opening the app.', bold=True)
line(7, 'Leaving a row blank', 'Fine, and it is the honest default. It stays archived and teaches nothing. Only rule on the ones you have a real view on.')
line(8, 'Duplicates', 'The Heads up column flags repeats. Where two rows are the same item, rule once and I will apply it to both.')
line(9, 'Sending it back', 'Save and return the file. Do not rename the tabs or move the columns - I read them by name.')

line(11, 'WHAT EACH CALL DOES', '', size=12, color='1F3864')
for i, (a, b) in enumerate([
    ('Good brief', 'Marks the brief approved. It stays in the archive; this records that it was worth sending.'),
    ('Not for me', 'Bins it AND writes the -1 with your reason. This is the one that teaches. Pick a WHY.'),
    ('Track this shift', 'Promotes the shift to active in the register, so it feeds future briefs.'),
    ('Same as another one', 'Folds this shift into an existing one. Put the row number in SAME AS ROW #.'),
    ('Not a shift', 'Rejects it with your reason. Teaches the detector what does not count.'),
    ('Yes, retire it', 'Confirms the shift is done. Where it says already retired, this just makes it official.'),
    ('No, still live', 'Keeps the shift active. The detector was wrong to call it quiet.'),
    ('Keep for good', 'Stamps the idea into the Library, permanently safe from any purge.'),
    ('Not worth keeping', 'Rejects the graduation with your reason.'),
    ('Got it', 'Acknowledges the investigation. No signal either way.'),
    ('Leave it', 'Explicitly no view. Same as blank.'),
], start=12):
    line(i, a, b)

line(24, 'WHAT EACH REASON MEANS', '', size=12, color='1F3864')
for i, (a, b) in enumerate([
    ('Wrong topic for me', 'The subject is not one you cover, however well made the card is.'),
    ('Wrong angle', 'Right subject, wrong take on it.'),
    ('Already covered', 'You have said this, and recently.'),
    ('Old news', 'True once. Not news by the time it surfaced.'),
    ('Too generic', 'Nothing here a reader could not guess.'),
    ('Not my voice', 'Would not survive being said out loud in your own words.'),
    ('Bad timing', 'Real and yours, just not now.'),
    ('Other', 'Anything else. Use NOTES and I will read it.'),
], start=25):
    line(i, a, b)

line(34, 'Why the reasons matter', 'They are clustered by code every Sunday. Refusing three cards for the same reason changes what gets assembled next, which is the whole point of filling this in. To date the system has received zero rejections, so it has learned nothing since July.')
rm.row_dimensions[34].height = 46

wb.remove(wb['Sheet'])
wb.save('content-archive-decisions.xlsx')
print('wrote content-archive-decisions.xlsx with', len(rows), 'rows')
