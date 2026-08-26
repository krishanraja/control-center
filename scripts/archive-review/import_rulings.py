"""Turn a filled-in content-archive-decisions.xlsx into an auditable SQL migration.

Mirrors what the live endpoints do, because these rows are 'archived' and
api/content-decisions/[id].ts refuses anything that is not 'pending' (409).

  Good brief          -> decision done, resolution action='approved'
  Not for me          -> decision dismissed + feedback_queue -1 with the reason
  Track this shift    -> shifts.status='active'            (api/shifts/[id] 'accept')
  Same as another one -> full merge: evidence + idea links move, source deleted
  Not a shift         -> reject + feedback -1 (the shift itself is left alone)
  Yes, retire it      -> shifts.status='retired'
  No, still live      -> shifts.status='active'
  Keep for good       -> content_ideas library_at + horizon evergreen + no expiry
  Got it              -> decision done, no signal
  Leave it / blank    -> untouched
"""
import sys, json, re
from openpyxl import load_workbook

REASON_CODE = {
    'Wrong topic for me': 'content_wrong_topic', 'Wrong angle': 'content_wrong_angle',
    'Already covered': 'content_already_covered', 'Old news': 'content_old_news',
    'Too generic': 'content_too_generic', 'Not my voice': 'content_not_my_voice',
    'Bad timing': 'content_bad_timing', 'Other': 'content_other',
}
REJECTS = {'Not for me', 'Not a shift', 'Not worth keeping'}
FIRST = 5

def q(s):
    return "'" + str(s).replace("'", "''") + "'" if s is not None else 'null'

def main(path, out):
    wb = load_workbook(path)
    ws = wb['Decide']
    rows, problems, sql = [], [], []
    by_num = {}
    for r in range(FIRST, ws.max_row + 1):
        did = ws.cell(row=r, column=14).value
        if not did:
            continue
        rec = dict(row=r, n=ws.cell(row=r, column=1).value, kind=ws.cell(row=r, column=3).value,
                   title=ws.cell(row=r, column=4).value, summary=ws.cell(row=r, column=5).value,
                   call=(ws.cell(row=r, column=10).value or '').strip(),
                   why=(ws.cell(row=r, column=11).value or '').strip(),
                   note=(ws.cell(row=r, column=12).value or '').strip(),
                   same_as=ws.cell(row=r, column=13).value, id=did)
        rows.append(rec); by_num[rec['n']] = rec

    ruled = [r for r in rows if r['call'] and r['call'] != 'Leave it']
    for r in ruled:
        if r['call'] in REJECTS and not r['why']:
            problems.append('Row %s: "%s" needs a WHY, or it teaches nothing.' % (r['n'], r['call']))
        if r['call'] == 'Same as another one':
            tgt = by_num.get(r['same_as'])
            if not tgt:
                problems.append('Row %s: merge needs a valid SAME AS ROW # (got %r).' % (r['n'], r['same_as']))
            elif tgt['kind'] not in ('New shift spotted', 'Shift going quiet'):
                problems.append('Row %s: can only merge into another shift, not "%s".' % (r['n'], tgt['kind']))

    if problems:
        print('\n'.join(problems)); print('\n%d problem(s) - nothing written.' % len(problems)); return 1

    sql.append('-- Generated from %s. %d of %d cards ruled on.' % (path, len(ruled), len(rows)))
    sql.append('begin;\n')
    for r in ruled:
        c, rid = r['call'], q(r['id'])
        sql.append('-- row %s  %s  %s' % (r['n'], r['kind'], (r['title'] or '')[:70]))
        if c in REJECTS:
            code = REASON_CODE.get(r['why'], 'content_other')
            res = {'action': 'reject', 'reason_code': code, 'reason_text': r['note'] or None,
                   'via': 'archive-review-2026-08'}
            sql.append("update content_decisions set status='dismissed', resolved_at=now(), "
                       "resolution=%s::jsonb where id=%s;" % (q(json.dumps(res)), rid))
            meta = {'kind': r['kind'], 'surface': 'archive_review',
                    'title': (r['title'] or '')[:300], 'text': (r['summary'] or '')[:1200]}
            sql.append("insert into feedback_queue (source_table, source_id, agent_id, original_agent, "
                       "original_item_id, vote, reason_code, reason_text, meta, status) values "
                       "('content_decisions', %s, 'cleo', 'cleo', %s, -1, %s, %s, %s::jsonb, 'pending');"
                       % (rid, rid, q(code), q(r['note'] or None), q(json.dumps(meta))))
        else:
            action = {'Good brief': 'approved', 'Track this shift': 'accept', 'Yes, retire it': 'retire',
                      'No, still live': 'keep_watching', 'Keep for good': 'library',
                      'Got it': 'acknowledged', 'Same as another one': 'merge'}[c]
            res = {'action': action, 'note': r['note'] or None, 'via': 'archive-review-2026-08'}
            sql.append("update content_decisions set status='done', resolved_at=now(), "
                       "resolution=%s::jsonb where id=%s;" % (q(json.dumps(res)), rid))

        if c == 'Track this shift' or c == 'No, still live':
            sql.append("update shifts set status='active' where id=(select ref from content_decisions where id=%s);" % rid)
        elif c == 'Yes, retire it':
            sql.append("update shifts set status='retired' where id=(select ref from content_decisions where id=%s);" % rid)
        elif c == 'Keep for good':
            sql.append("update content_ideas set library_at=now(), horizon='evergreen', expires_at=null "
                       "where id=(select ref from content_decisions where id=%s);" % rid)
        elif c == 'Same as another one':
            tgt = by_num[r['same_as']]
            sql.append("-- merge into row %s: %s" % (tgt['n'], (tgt['title'] or '')[:60]))
            src = "(select ref from content_decisions where id=%s)" % rid
            dst = "(select ref from content_decisions where id=%s)" % q(tgt['id'])
            sql.append("insert into shift_evidence (shift_id, occurred_on, headline, source, url, "
                       "provenance, week_label, citable, quarantine_reason) select %s, occurred_on, "
                       "headline, source, url, provenance, week_label, citable, quarantine_reason "
                       "from shift_evidence where shift_id=%s on conflict (shift_id, occurred_on, headline) "
                       "do nothing;" % (dst, src))
            sql.append("delete from shift_evidence where shift_id=%s;" % src)
            sql.append("update content_ideas set shift_id=%s where shift_id=%s;" % (dst, src))
            sql.append("delete from shifts where id=%s;" % src)
        sql.append('')

    sql.append("insert into audit_log (event_type, actor, target, details) values ('content_archive_review', "
               "'krish', 'content_decisions', %s);"
               % q(json.dumps({'ruled': len(ruled), 'of': len(rows), 'source': path})))
    sql.append('\ncommit;')
    open(out, 'w').write('\n'.join(sql) + '\n')
    print('%d of %d ruled on -> %s' % (len(ruled), len(rows), out))
    print('  rejections that will teach:', sum(1 for r in ruled if r['call'] in REJECTS))
    return 0

if __name__ == '__main__':
    sys.exit(main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else 'rulings.sql'))
